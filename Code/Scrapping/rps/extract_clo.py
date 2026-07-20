"""
Extract CLO (Course Learning Outcome) data from Telkom University RPS
files in .xlsx format.

Grabs 4 fields per weekly row: ID CLO, Deskripsi (Sub-CLO), Indikator,
Materi. Works even if the sheet name or exact merged-cell widths differ
between matkul, since columns are located dynamically by header text
instead of hardcoded column letters.

Usage:
    Single file : python3 extract_clo_xlsx.py "RPS - Matkul.xlsx"
    Batch       : python3 extract_clo_xlsx.py root_folder output.xlsx
"""
import os
import sys
import re
import openpyxl
import pandas as pd

HEADER_ALIASES = {
    "id_clo": ["ID CLO"],
    "deskripsi": ["DESKRIPSI SUB CLO", "DESKRIPSI SUB-CLO", "DESKRIPSI"],
    "indikator": ["INDIKATOR KETERCAPAIAN CLO", "INDIKATOR"],
    "materi": ["MATERI"],
}


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


def find_data_sheet(wb):
    """Pick the sheet that actually contains the weekly CLO table."""
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows(max_row=min(ws.max_row, 200)):
            for cell in row:
                if norm(cell.value) == "MINGGU KE-":
                    return ws
    return None


def find_header_row_and_cols(ws):
    for r in range(1, min(ws.max_row, 200) + 1):
        row_vals = {c: norm(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)}
        if "MINGGU KE-" in row_vals.values():
            cols = {}
            for key, aliases in HEADER_ALIASES.items():
                aliases_norm = [norm(a) for a in aliases]
                for c, v in row_vals.items():
                    if v in aliases_norm:
                        cols[key] = c
                        break
            return r, cols
    return None, {}


def find_mata_kuliah(ws):
    for row in ws.iter_rows(max_row=min(ws.max_row, 30)):
        for cell in row:
            if norm(cell.value) == "NAMA MK":
                val = ws.cell(row=cell.row + 1, column=cell.column).value
                if val:
                    return str(val).strip()
    return None


def clean_text(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def process_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_data_sheet(wb)
    if ws is None:
        raise ValueError("Tabel CLO (baris 'MINGGU KE-') tidak ditemukan di sheet manapun")

    header_row, cols = find_header_row_and_cols(ws)
    missing = [k for k in HEADER_ALIASES if k not in cols]
    if missing:
        raise ValueError(f"Kolom tidak ditemukan: {missing}")

    matkul = find_mata_kuliah(ws) or os.path.splitext(os.path.basename(path))[0]

    records = []
    for r in range(header_row + 1, ws.max_row + 1):
        id_clo = ws.cell(row=r, column=cols["id_clo"]).value
        if not id_clo or not re.match(r"^\s*CLO\s*\d+\s*$", str(id_clo)):
            continue
        records.append({
            "Mata Kuliah": matkul,
            "ID CLO": clean_text(id_clo),
            "Deskripsi": clean_text(ws.cell(row=r, column=cols["deskripsi"]).value),
            "Indikator": clean_text(ws.cell(row=r, column=cols["indikator"]).value),
            "Materi": clean_text(ws.cell(row=r, column=cols["materi"]).value),
            "source_file": os.path.basename(path),
        })
    return records


def find_xlsx_files(root):
    paths = []
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            if fn.lower().endswith(".xlsx") and not fn.startswith("~$"):
                paths.append(os.path.join(dirpath, fn))
    return sorted(paths)


def run_batch(root, output_path):
    files = find_xlsx_files(root)
    if not files:
        print(f"[WARN] Tidak ada file .xlsx ditemukan di: {root}", file=sys.stderr)
        return

    all_records = []
    failed = []
    for path in files:
        try:
            records = process_xlsx(path)
            if not records:
                print(f"[WARN] 0 baris CLO dari: {path}", file=sys.stderr)
            all_records.extend(records)
        except Exception as e:
            failed.append((path, str(e)))
            print(f"[ERROR] Gagal proses {path}: {e}", file=sys.stderr)

    df = pd.DataFrame(all_records)
    df.to_excel(output_path, index=False)

    print(f"Selesai. {len(files)} file diproses, {len(df)} baris CLO dihasilkan.")
    print(f"Output: {output_path}")
    if failed:
        print(f"\n{len(failed)} file GAGAL diproses:")
        for p, err in failed:
            print(f"  - {p}: {err}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  Satu file : python3 extract_clo_xlsx.py <file.xlsx>")
        print("  Batch     : python3 extract_clo_xlsx.py <root_folder> <output.xlsx>")
        sys.exit(1)

    target = sys.argv[1]

    if os.path.isdir(target):
        output_path = sys.argv[2] if len(sys.argv) > 2 else "clo_output.xlsx"
        run_batch(target, output_path)
    else:
        records = process_xlsx(target)
        df = pd.DataFrame(records)
        pd.set_option("display.max_colwidth", 80)
        print(df.to_string())
        print(f"\nTotal CLO rows: {len(df)}")