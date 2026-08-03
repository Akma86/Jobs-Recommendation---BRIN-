# -*- coding: utf-8 -*-
"""
HUMAN/EXPERT VALIDATION TEMPLATE - generates a spreadsheet of (course/cert
evidence, job) pairs for a human (you, or ideally your advisor) to rate
manually 0-5 on "does this job genuinely fit this evidence", then computes
correlation between those human ratings and the pipeline's final_score.

This mirrors exactly what Reyhan did for CareerSync (his slide showed
r=0.896 correlation between the model's cosine similarity and expert
judgment) - having the equivalent number for your pipeline is one of the
strongest pieces of evidence you can bring to a thesis defense, because it
answers the question "does this actually make sense to a human" directly,
rather than through a proxy metric.

STRATIFIED SAMPLING: pulls from final_recommendations_full.csv across the
FULL score range (top, middle, bottom), not just the top-N. Rating only
top recommendations tells you nothing about whether the pipeline correctly
identifies BAD matches too - a system that's "correct" only ever gets
tested on the easy top hits looks better than it is.

REQUIREMENTS:
  pip install openpyxl pandas scipy --break-system-packages

USAGE:
  # Step 1 - generate the rating template:
  python human_validation_template.py generate --n_per_stratum 5

  # Step 2 - open human_validation_template.xlsx, fill in the yellow
  # "Rating (0-5)" column by hand (or send it to your advisor)

  # Step 3 - after filling it in, compute the correlation:
  python human_validation_template.py analyze --file human_validation_template.xlsx
"""

import argparse
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from scipy.stats import spearmanr, pearsonr

FONT_NAME = "Arial"


def stratified_sample(final_ranking, n_per_stratum=5, seed=42):
    df = final_ranking.sort_values("final_score", ascending=False).reset_index(drop=True)
    n = len(df)
    if n < 3 * n_per_stratum:
        print(f"WARNING: only {n} scored jobs available - sampling what's there "
              f"instead of the requested {3*n_per_stratum} (top/mid/bottom).")
        return df.sample(min(n, 3 * n_per_stratum), random_state=seed)

    third = n // 3
    top = df.iloc[:third].sample(min(n_per_stratum, third), random_state=seed)
    mid = df.iloc[third:2*third].sample(min(n_per_stratum, third), random_state=seed)
    bottom = df.iloc[2*third:].sample(min(n_per_stratum, n - 2*third), random_state=seed)
    sample = pd.concat([top, mid, bottom]).sample(frac=1, random_state=seed)  # shuffle so rater can't tell stratum
    return sample.reset_index(drop=True)


def generate_template(n_per_stratum, out_path="human_validation_template.xlsx"):
    final_ranking = pd.read_csv("final_recommendations_full.csv")
    sample = stratified_sample(final_ranking, n_per_stratum)

    wb = Workbook()
    ws = wb.active
    ws.title = "Validasi"

    # legend
    ws["A1"] = "PETUNJUK: isi kolom kuning 'Rating (0-5)' untuk tiap baris — seberapa cocok pekerjaan ini dengan bukti kompetensi (matkul/sertifikat) yang tertera. 0 = sama sekali tidak cocok, 5 = sangat cocok."
    ws["A1"].font = Font(name=FONT_NAME, italic=True, size=10)
    ws.merge_cells("A1:F1")

    headers = ["job_id", "job_title", "job_company", "explanation (bukti dari pipeline)",
               "final_score (pipeline)", "Rating (0-5)"]
    header_row = 3
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = Font(name=FONT_NAME, bold=True)

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for i, (_, row) in enumerate(sample.iterrows(), start=header_row + 1):
        ws.cell(row=i, column=1, value=row["job_id"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=2, value=row["job_title"]).font = Font(name=FONT_NAME)
        ws.cell(row=i, column=3, value=row.get("job_company", "")).font = Font(name=FONT_NAME)
        expl_cell = ws.cell(row=i, column=4, value=row.get("explanation", ""))
        expl_cell.font = Font(name=FONT_NAME)
        expl_cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=5, value=round(row["final_score"], 3)).font = Font(name=FONT_NAME)
        rating_cell = ws.cell(row=i, column=6, value=None)
        rating_cell.fill = yellow
        rating_cell.font = Font(name=FONT_NAME)

    # example row appended below, clearly marked, showing expected format
    example_row = header_row + len(sample) + 2
    ws.cell(row=example_row, column=1, value="CONTOH (bukan data asli):").font = Font(name=FONT_NAME, italic=True)
    ws.cell(row=example_row + 1, column=2, value="Data Analyst").font = Font(name=FONT_NAME)
    ws.cell(row=example_row + 1, column=6, value=4).font = Font(name=FONT_NAME)
    ws.cell(row=example_row + 1, column=6).fill = yellow

    widths = [16, 30, 20, 60, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"Saved: {out_path} ({len(sample)} pairs to rate)")
    print("Fill in the yellow 'Rating (0-5)' column, then run:")
    print(f"  python human_validation_template.py analyze --file {out_path}")


def analyze(file_path):
    df = pd.read_excel(file_path, sheet_name="Validasi", header=2)  # header is on row 3 (0-indexed row 2)
    df = df.dropna(subset=["Rating (0-5)", "job_id", "final_score (pipeline)"])
    df = df[pd.to_numeric(df["Rating (0-5)"], errors="coerce").notna()]
    df = df[pd.to_numeric(df["final_score (pipeline)"], errors="coerce").notna()]

    if len(df) < 5:
        print(f"Only {len(df)} rated rows found - need at least a handful for a meaningful correlation. "
              "Make sure you filled in the yellow column and saved the file.")
        return

    ratings = df["Rating (0-5)"].astype(float)
    scores = df["final_score (pipeline)"].astype(float)

    rho, p_spear = spearmanr(ratings, scores)
    r, p_pear = pearsonr(ratings, scores)

    print(f"=== Validasi Human Judgment (n={len(df)}) ===")
    print(f"Spearman rho = {rho:.3f}  (p={p_spear:.4f})")
    print(f"Pearson r    = {r:.3f}  (p={p_pear:.4f})")
    print()
    if rho > 0.7:
        print("-> Korelasi KUAT: rating manusia dan skor pipeline sejalan. Kabar baik untuk defense.")
    elif rho > 0.4:
        print("-> Korelasi SEDANG: ada keselarasan, tapi masih ada disagreement yang perlu dicek manual "
              "(lihat baris dengan selisih rating vs final_score terbesar).")
    else:
        print("-> Korelasi LEMAH: pipeline dan penilaian manusia sering gak sejalan. Ini temuan penting - "
              "cek baris-baris dengan disagreement terbesar untuk paham polanya (skor terlalu bergantung "
              "1 sinyal? bobot grade/cert perlu di-tune?).")

    df["disagreement"] = (df["Rating (0-5)"] - (df["final_score (pipeline)"] / df["final_score (pipeline)"].max() * 5)).abs()
    print("\n--- Top 3 disagreement terbesar (worth dicek manual) ---")
    print(df.sort_values("disagreement", ascending=False).head(3)[
        ["job_title", "Rating (0-5)", "final_score (pipeline)"]].to_string(index=False))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest="command", required=True)

    gen_p = sub.add_parser("generate")
    gen_p.add_argument("--n_per_stratum", type=int, default=5)
    gen_p.add_argument("--out", default="human_validation_template.xlsx")

    an_p = sub.add_parser("analyze")
    an_p.add_argument("--file", default="human_validation_template.xlsx")

    args = parser.parse_args()
    if args.command == "generate":
        generate_template(args.n_per_stratum, args.out)
    else:
        analyze(args.file)
