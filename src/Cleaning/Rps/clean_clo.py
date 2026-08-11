import re
import openpyxl

INPUT_PATH = "output.xlsx"
OUTPUT_PATH = "output_split.xlsx"

# Tag Sub-CLO seperti:
# [CLO 1-1.1]
# [CLO 1-Sub CLO 04]
# [SubCLO-03]
# [CLO 1]
TAG_PATTERN = re.compile(r"\[[^\]\n]{1,60}\]")

# Label field yang mungkin muncul di dalam satu chunk sub-CLO.
# Urutan biasanya Hasil -> Indikator -> Materi, tapi gak semua chunk
# punya ketiganya lengkap / urut sama persis, jadi dicari per-label.
LABEL_PATTERN = re.compile(r"(Hasil|Indikator|Materi)\s*:", re.IGNORECASE)


def extract_fields(chunk):
    """
    Return (hasil, indikator, materi) dari satu chunk sub-CLO.
    Cara kerja: cari semua posisi label (Hasil:/Indikator:/Materi:) di
    dalam chunk, lalu teks tiap field = teks setelah labelnya sampai
    sebelum label berikutnya (atau sampai akhir chunk kalau dia label
    terakhir). Kalau suatu label gak ada di chunk, fieldnya dikosongin.
    """
    result = {"hasil": "", "indikator": "", "materi": ""}
    if not chunk:
        return result["hasil"], result["indikator"], result["materi"]

    matches = list(LABEL_PATTERN.finditer(chunk))
    if not matches:
        return result["hasil"], result["indikator"], result["materi"]

    for i, m in enumerate(matches):
        label = m.group(1).lower()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(chunk)
        value = chunk[start:end].strip()
        value = value.rstrip("|").strip()
        if label in result and not result[label]:
            result[label] = value
        elif label in result:
            result[label] = (result[label] + " " + value).strip()

    return result["hasil"], result["indikator"], result["materi"]


def split_full_clo_text(text, fallback_clo_id):
    if not text:
        return []
    matches = list(TAG_PATTERN.finditer(text))
    if not matches:
        return [(fallback_clo_id, text.strip())]

    results = []
    prefix = text[:matches[0].start()].strip()
    if prefix:
        results.append((fallback_clo_id, prefix))

    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        chunk = text[start:end].strip()
        if not chunk:
            continue
        tag_text = m.group(0)[1:-1].strip()
        results.append((tag_text, chunk))

    return results


def main():
    wb = openpyxl.load_workbook(INPUT_PATH)
    # ambil sheet pertama supaya tidak tergantung nama
    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in ws[1]]
    col_idx = {name: i for i, name in enumerate(headers)}

    required = [
        "Mata Kuliah",
        "CLO-ID",
        "full_clo_text"
    ]
    for col in required:
        if col not in col_idx:
            raise ValueError(f"Kolom '{col}' tidak ditemukan")

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Split_CLO"

    out_header = [
        "Mata Kuliah",
        "CLO-ID",
        "Hasil",
        "Indikator",
        "full_clo_text",
        "source_file"
    ]
    out_ws.append(out_header)

    total_sub = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        mata_kuliah = row[col_idx["Mata Kuliah"]]
        clo_id = row[col_idx["CLO-ID"]]
        full_text = row[col_idx["full_clo_text"]]
        source_file = (
            row[col_idx["source_file"]]
            if "source_file" in col_idx
            else ""
        )

        # tag_text (Sub-CLO) tetap dipakai secara internal untuk logic
        # pemotongan chunk, tapi gak diikutkan sebagai kolom output.
        sub_entries = split_full_clo_text(full_text, clo_id)

        for _tag_text, chunk in sub_entries:
            hasil, indikator, _materi = extract_fields(chunk)
            out_ws.append([
                mata_kuliah,
                clo_id,
                hasil,
                indikator,
                chunk,
                source_file
            ])
            total_sub += 1

    column_widths = {
        "A": 35,
        "B": 15,
        "C": 60,
        "D": 60,
        "E": 100,
        "F": 40,
    }
    for col, width in column_widths.items():
        out_ws.column_dimensions[col].width = width

    out_wb.save(OUTPUT_PATH)
    print(
        f"Selesai. Total baris asli: {ws.max_row - 1}, "
        f"total sub-CLO hasil split: {total_sub}"
    )
    print(f"Output: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()